import io, zipfile, datetime
import pandas as pd
import streamlit as st
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.page import PageMargins

# ç”»é¢è¨­å®š
st.set_page_config(page_title="ã„ã‚ã¨ç™ºæ³¨ç®¡ç†", layout="centered")

# å…±é€šã‚¹ã‚¿ã‚¤ãƒ«
TITLE = "ã„ã‚ã¨ç™ºæ³¨ç®¡ç†"
LEFT_HEADER_FONT   = Font(name="ï¼­ï¼³ ã‚´ã‚·ãƒƒã‚¯", size=28, bold=True)
CENTER_HEADER_FONT = Font(name="ï¼­ï¼³ ã‚´ã‚·ãƒƒã‚¯", size=26, bold=True)
RIGHT_HEADER_FONT  = Font(name="ï¼­ï¼³ ã‚´ã‚·ãƒƒã‚¯", size=24, bold=True)
BODY_FONT          = Font(name="ï¼­ï¼³ ã‚´ã‚·ãƒƒã‚¯", size=22)
THIN               = Side(border_style="thin", color="000000")

def style_sheet(ws):
    # æœ¬æ–‡ãƒ•ã‚©ãƒ³ãƒˆãƒ»ç½«ç·šãƒ»è¡Œé«˜
    for row in ws.iter_rows(min_row=6):
        for c in row:
            c.font = BODY_FONT
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    for i in range(1, ws.max_row + 1):
        ws.row_dimensions[i].height = 30
    # A4æ¨ª
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.5)

def set_header(ws, supplier, facility_label):
    # å·¦ï¼šä»•å…¥å…ˆ å¾¡ä¸­ï¼ˆA3:B3ï¼‰
    ws.merge_cells("A3:B3")
    ws["A3"] = f"{supplier}ã€€å¾¡ä¸­"
    ws["A3"].font = LEFT_HEADER_FONT
    ws["A3"].alignment = Alignment(horizontal="left", vertical="bottom")
    # ä¸­å¤®ï¼šæ³¨æ–‡æ›¸ï¼ˆæ–½è¨­åï¼‰ï¼ˆB1:F1ï¼‰
    ws.merge_cells("B1:F1")
    ws["B1"] = f"æ³¨æ–‡æ›¸ï¼ˆ{facility_label}ï¼‰"
    ws["B1"].font = CENTER_HEADER_FONT
    ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
    # å³ï¼š(æœ‰) ãƒãƒ¼ãƒˆãƒŸãƒ¼ãƒ«ï¼ˆM2:O2ï¼‰
    ws.merge_cells("M2:O2")
    ws["M2"] = "(æœ‰) ãƒãƒ¼ãƒˆãƒŸãƒ¼ãƒ«"
    ws["M2"].font = RIGHT_HEADER_FONT
    ws["M2"].alignment = Alignment(horizontal="right", vertical="center")
    # è¦‹å‡ºã—è¿‘è¾ºã®è¡Œé«˜
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 35
    ws.row_dimensions[3].height = 35

def ensure_columns(df, cols):
    for c in cols:
        if c not in df.columns:
            df[c] = ""
    return df

def forward_fill_cols(df, cols):
    for c in cols:
        if c in df.columns:
            df[c] = df[c].ffill()
    return df

def read_excel_flexible(file_like, header_row):
    """BytesIOã‹ã‚‰ã§ã‚‚ç¢ºå®Ÿã«èª­ã‚€ï¼ˆ2è¡Œãƒ˜ãƒƒãƒ€ãƒ¼å¯¾å¿œï¼‰ã€‚header_rowã¯1å§‹ã¾ã‚Šã€‚"""
    hdr = max(0, header_row - 1)
    try:
        # ğŸ”¹ 2è¡Œãƒ˜ãƒƒãƒ€ãƒ¼ã‚’ä¸€ä½“åŒ–ã—ã¦èª­ã¿è¾¼ã¿
        df = pd.read_excel(file_like, header=[hdr, hdr + 1])
        df.columns = [
            ''.join([str(c) for c in col if str(c) != 'nan']).replace('Unnamed: ', '').strip()
            for col in df.columns
        ]
    except Exception:
        # ğŸ”¹ ãƒ•ã‚©ãƒ¼ãƒ«ãƒãƒƒã‚¯ï¼ˆ1è¡Œãƒ˜ãƒƒãƒ€ãƒ¼ï¼‰
        df = pd.read_excel(file_like, header=hdr)
        df.columns = df.columns.astype(str).str.strip().str.replace("\n", "", regex=False)
    return df


def to_excel_bytes(df, startrow=0):
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, startrow=startrow)
    bio.seek(0)
    return bio.getvalue()

def output_order_excels_zip(df, facility):
    # åˆ—åã‚†ã‚‰ãã‚’å¸å
    rename_map = {}
    for c in df.columns:
        cc = str(c)
        if "ä½¿ç”¨æ—¥" in cc: rename_map[c] = "ä½¿ç”¨æ—¥"
        if "ä»•å…¥å…ˆ" in cc: rename_map[c] = "ä»•å…¥å…ˆ"
        if "é£Ÿå“å" in cc: rename_map[c] = "é£Ÿå“å"
        if "å˜ä½"   in cc and "ãƒ¦ãƒ‹" not in cc: rename_map[c] = "å˜ä½"
        if "å…¥æ‰€è€…" in cc and "ãƒ¦" not in cc: rename_map[c] = "å…¥æ‰€è€…"
        if "è·å“¡"   in cc: rename_map[c] = "è·å“¡"
        if "ãƒ¦" in cc and "å…¥æ‰€è€…" in cc: rename_map[c] = "ãƒ¦ãƒ¼ãƒã‚¦ã‚¹å…¥æ‰€è€…"
        if "å‚™è€ƒ"   in cc: rename_map[c] = "å‚™è€ƒæ¬„"
        if "ç´å“æ™‚é–“" in cc: rename_map[c] = "ç´å“æ™‚é–“"
        if "æ¤œå" in cc: rename_map[c] = "æ¤œåè€…"
        if "é®®åº¦" in cc: rename_map[c] = "é®®åº¦"
        if "å“æ¸©" in cc: rename_map[c] = "å“æ¸©"
        if "ç•°ç‰©" in cc: rename_map[c] = "ç•°ç‰©"
        if "åŒ…è£…" in cc: rename_map[c] = "åŒ…è£…"
        if "æœŸé™" in cc: rename_map[c] = "æœŸé™"
    df = df.rename(columns=rename_map)

    # æ¬ æè£œå®Œã¨å‹
    df = forward_fill_cols(df, ["ä½¿ç”¨æ—¥","ä»•å…¥å…ˆ","é£Ÿå“å"])
    if facility == "ã„ã‚ã¨":
        for c in ["å…¥æ‰€è€…","è·å“¡"]:
            if c in df.columns:
                df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    else:
        if "ãƒ¦ãƒ¼ãƒã‚¦ã‚¹å…¥æ‰€è€…" in df.columns:
            df["ãƒ¦ãƒ¼ãƒã‚¦ã‚¹å…¥æ‰€è€…"] = pd.to_numeric(df["ãƒ¦ãƒ¼ãƒã‚¦ã‚¹å…¥æ‰€è€…"], errors="coerce").fillna(0)

    if "ä»•å…¥å…ˆ" not in df.columns:
        raise ValueError("ã€Œä»•å…¥å…ˆã€åˆ—ãŒè¦‹ã¤ã‹ã‚Šã¾ã›ã‚“ã€‚ãƒ˜ãƒƒãƒ€ãƒ¼è¡Œã®æŒ‡å®šã‚’è¦‹ç›´ã—ã¦ãã ã•ã„ã€‚")

    suppliers = df["ä»•å…¥å…ˆ"].dropna().unique()

    # å›ºå®šã®åˆ—é †ï¼ˆæ–½è¨­åˆ¥ï¼‰
    if facility == "ã„ã‚ã¨":
        keep_cols = ["ä½¿ç”¨æ—¥","é£Ÿå“å","å…¥æ‰€è€…","å˜ä½","è·å“¡",
                     "é®®åº¦","å“æ¸©","ç•°ç‰©","åŒ…è£…","æœŸé™","å‚™è€ƒæ¬„","ç´å“æ™‚é–“","æ¤œåè€…"]
        group_by = ["ä½¿ç”¨æ—¥","é£Ÿå“å","å˜ä½"]
        agg = {"å…¥æ‰€è€…":"sum", "è·å“¡":"sum"}
        facility_label = "ä»‹è­·è€äººç¦ç¥‰æ–½è¨­ã„ã‚ã¨"
    else:
        keep_cols = ["ä½¿ç”¨æ—¥","é£Ÿå“å","ãƒ¦ãƒ¼ãƒã‚¦ã‚¹å…¥æ‰€è€…","å˜ä½",
                     "é®®åº¦","å“æ¸©","ç•°ç‰©","åŒ…è£…","æœŸé™","å‚™è€ƒæ¬„","ç´å“æ™‚é–“","æ¤œåè€…"]
        group_by = ["ä½¿ç”¨æ—¥","é£Ÿå“å","å˜ä½"]
        agg = {"ãƒ¦ãƒ¼ãƒã‚¦ã‚¹å…¥æ‰€è€…":"sum"}
        facility_label = "ãƒ¦ãƒ¼ãƒã‚¦ã‚¹ã„ã‚ã¨"

    # ZIPã«ã¾ã¨ã‚ã‚‹
    zip_bytes = io.BytesIO()
    with zipfile.ZipFile(zip_bytes, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for supplier in suppliers:
            sub = df[df["ä»•å…¥å…ˆ"] == supplier].copy()
            present_keys = [k for k in group_by if k in sub.columns]
            sub = sub.groupby(present_keys, as_index=False).agg(agg)
            sub = ensure_columns(sub, keep_cols)
            sub = sub[keep_cols]

            wb = Workbook()
            ws = wb.active

            # ãƒ†ãƒ¼ãƒ–ãƒ«å‡ºåŠ›ï¼ˆ6è¡Œç›®ã‹ã‚‰ï¼‰
            tmp = io.BytesIO()
            with pd.ExcelWriter(tmp, engine="openpyxl") as writer:
                sub.to_excel(writer, index=False, startrow=5)
            tmp.seek(0)
            tmp_wb = load_workbook(tmp)
            tmp_ws = tmp_wb.active
            for r in tmp_ws.iter_rows(values_only=True):
                ws.append(list(r))

            style_sheet(ws)
            set_header(ws, str(supplier), facility_label)

            safe = str(supplier).replace("/", "_").replace("\\", "_")
            out_name = f"æ³¨æ–‡æ›¸_{safe}_{facility}.xlsx"
            out_bio = io.BytesIO()
            wb.save(out_bio)
            out_bio.seek(0)
            zf.writestr(out_name, out_bio.read())

    zip_bytes.seek(0)
    return zip_bytes.getvalue()

# ================= UI =================
st.title(TITLE)
st.caption("ãƒ–ãƒ©ã‚¦ã‚¶ã ã‘ã§ã€æ¤œåç°¿ã®åŠ å·¥ â†’ ä»•å…¥å…ˆåˆ¥æ³¨æ–‡æ›¸ï¼ˆã„ã‚ã¨ï¼ãƒ¦ãƒ¼ãƒã‚¦ã‚¹ï¼‰ã€ã‚’ä½œæˆã—ã¾ã™ã€‚")

# ---------- STEP 1ï¼šæ¤œåç°¿ã®åŠ å·¥ ----------
with st.expander("STEP 1ï¼šæ¤œåç°¿ã®åŠ å·¥ï¼ˆç©ºæ¬„è£œå®Œä»˜ãï¼‰", expanded=True):
    uploaded_raw = st.file_uploader("æ¤œåè¨˜éŒ²ç°¿ï¼ˆåŸæœ¬ .xlsxï¼‰ã‚’ã‚¢ãƒƒãƒ—ãƒ­ãƒ¼ãƒ‰", type=["xlsx"], key="raw")
    header_row = st.number_input("ãƒ˜ãƒƒãƒ€ãƒ¼ï¼ˆè¦‹å‡ºã—ï¼‰ã®è¡Œç•ªå·ï¼ˆ1å§‹ã¾ã‚Šï¼‰", min_value=1, value=1, step=1)
    cols_to_ffill = st.multiselect(
        "ä¸‹æ–¹å‘ã«ã‚³ãƒ”ãƒ¼ï¼ˆç©ºæ¬„è£œå®Œï¼‰ã™ã‚‹åˆ—å",
        options=["ç´å“æ—¥","ä½¿ç”¨æ—¥","æœæ˜¼å¤•","ä»•å…¥å…ˆ","é£Ÿå“å"],
        default=["ç´å“æ—¥","ä½¿ç”¨æ—¥","æœæ˜¼å¤•","ä»•å…¥å…ˆ"]
    )

    if st.button("åŠ å·¥ã™ã‚‹ â–¶", use_container_width=True, disabled=(uploaded_raw is None)):
        try:
            # â†â†â† é‡è¦ï¼šBytesIOã§â€œç¢ºå®Ÿã«â€èª­ã¿è¾¼ã‚€
            raw_bytes = uploaded_raw.read()
            raw_excel = BytesIO(raw_bytes)

            df_raw = read_excel_flexible(raw_excel, header_row)
            df_proc = forward_fill_cols(df_raw, cols_to_ffill)

            ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            st.success("âœ… åŠ å·¥ãŒå®Œäº†ã—ã¾ã—ãŸã€‚ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰ã—ã¦ãã ã•ã„ã€‚")
            st.download_button(
                label="åŠ å·¥æ¸ˆãƒ•ã‚¡ã‚¤ãƒ«ã‚’ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰",
                data=to_excel_bytes(df_proc, startrow=0),
                file_name=f"æ¤œåç°¿_åŠ å·¥æ¸ˆ_ç©ºæ¬„è£œå®Œæ¸ˆã¿_{ts}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        except Exception as e:
            st.error(f"âŒ åŠ å·¥ä¸­ã«ã‚¨ãƒ©ãƒ¼ï¼š{e}")

st.markdown("---")

# ---------- STEP 2ï¼šä»•å…¥å…ˆåˆ¥ æ³¨æ–‡æ›¸ ----------
with st.expander("STEP 2ï¼šä»•å…¥å…ˆåˆ¥ æ³¨æ–‡æ›¸ã‚’ä½œæˆï¼ˆZIPï¼‰", expanded=True):
    uploaded_proc = st.file_uploader("åŠ å·¥æ¸ˆãƒ•ã‚¡ã‚¤ãƒ«ï¼ˆ.xlsxï¼‰ã‚’ã‚¢ãƒƒãƒ—ãƒ­ãƒ¼ãƒ‰", type=["xlsx"], key="proc")
    facility = st.radio("æ–½è¨­ã‚’é¸æŠ", options=["ã„ã‚ã¨","ãƒ¦ãƒ¼ãƒã‚¦ã‚¹"], horizontal=True)
    st.caption("å‡ºåŠ›ä»•æ§˜ï¼šA4æ¨ªï¼MSã‚´ã‚·ãƒƒã‚¯22ptï¼è¡Œé«˜30ï¼ç´°ç½«ç·šï¼ãƒ˜ãƒƒãƒ€ãƒ¼ï¼ˆå·¦ï¼šå¾¡ä¸­ï¼ä¸­å¤®ï¼šæ–½è¨­åï¼å³ï¼š(æœ‰)ãƒãƒ¼ãƒˆãƒŸãƒ¼ãƒ«ï¼‰ï¼æ¤œåè€…åˆ—ã‚ã‚Š")

    if st.button("æ³¨æ–‡æ›¸ã‚’ä½œæˆ â–¶", use_container_width=True, disabled=(uploaded_proc is None)):
        try:
            proc_bytes = uploaded_proc.read()
            proc_excel = BytesIO(proc_bytes)

            df2 = pd.read_excel(proc_excel, header=0)
            zip_data = output_order_excels_zip(df2, facility=facility)
            ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            st.success("âœ… ä»•å…¥å…ˆåˆ¥ã®æ³¨æ–‡æ›¸ã‚’ZIPã§ç”¨æ„ã—ã¾ã—ãŸã€‚ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰ã—ã¦ãã ã•ã„ã€‚")
            st.download_button(
                label="æ³¨æ–‡æ›¸ZIPã‚’ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰",
                data=zip_data,
                file_name=f"æ³¨æ–‡æ›¸_{facility}_{ts}.zip",
                mime="application/zip",
                use_container_width=True
            )
        except Exception as e:
            st.error(f"âŒ ä½œæˆä¸­ã«ã‚¨ãƒ©ãƒ¼ï¼š{e}")



